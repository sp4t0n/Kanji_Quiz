import tkinter as tk
from tkinter import messagebox, simpledialog, ttk
from ttkbootstrap import Style
from openpyxl import Workbook, load_workbook
import os
import random

DATA_FILE = "quiz_data.xlsx"

class QuizApp:
    
    # Initialization and Setup
    # Queste sono le funzioni che inizializzano l'applicazione e impostano l'interfaccia utente.
    
    def __init__(self, root):
        self.root = root
        self.root.title("Kanji Quiz")
        self.quiz_categories = []
        self.selected_categories = []
        self.current_category = ""
        self.current_quiz = None
        self.score = 0
        self.correct_answers = 0
        self.total_questions = 0
        self.shown_quizzes = {}
        self.quiz_data = {}
        self.quiz_direction = 'kanji to meaning' 

        self.category_var = tk.StringVar(value=self.selected_categories)
        self.category_listbox = tk.Listbox(root, listvariable=self.category_var, selectmode=tk.MULTIPLE, font=('Arial', 12))

        # Adding Scrollbar
        self.category_scrollbar = ttk.Scrollbar(root, orient="vertical", command=self.category_listbox.yview)
        self.category_listbox.configure(yscrollcommand=self.category_scrollbar.set)

        self.switch_mode_button = ttk.Button(root, text='Cambia modalità', style='warning.TButton', command=self.switch_mode)
        
        self.question_label = ttk.Label(root, font=('Arial', 24))

        # self.answer_entry = ttk.Entry(root, font=('Arial', 18))
        
        self.option1_button = tk.Button(self.root, text="", font=('Arial', 14), command=lambda: self.check_answer(1))
        
        self.option2_button = tk.Button(self.root, text="", font=('Arial', 14), command=lambda: self.check_answer(2))
        
        self.option3_button = tk.Button(self.root, text="", font=('Arial', 14), command=lambda: self.check_answer(3))
        
        self.submit_button = ttk.Button(root, text='Invia', style='danger.TButton', command=self.check_answer, state=tk.DISABLED)

        self.next_button = ttk.Button(root, text='Prossima domanda', style='success.TButton', command=self.next_question, state=tk.DISABLED)

        self.score_label = tk.Label(self.root, text="Punteggio: 0/0", font=('Arial', 14))
 
        self.show_answers_button = tk.Button(self.root, text="Mostra risposte", command=self.show_answers)
        
        self.add_category_button = ttk.Button(root, text='Aggiungi Categoria', style='info.TButton', command=self.open_add_category_window)

        self.edit_category_button = ttk.Button(root, text='Modifica Categoria', style='info.TButton', command=self.open_edit_category_window)

        self.add_quiz_button = ttk.Button(root, text='  Aggiungi Quiz  ', style='info.TButton', command=self.open_add_quiz_window)

        self.edit_quiz_button = ttk.Button(root, text='  Modifica Quiz  ', style='info.TButton', command=self.open_edit_quiz_window)

        self.category_listbox.bind('<<ListboxSelect>>', self.load_quiz)

        self.load_quiz_data()

        self.create_layout()


    def create_layout(self):
        self.add_category_button.grid(row=0, column=0, pady=10, padx=10)
        self.edit_category_button.grid(row=1, column=0, pady=10, padx=10)
        self.category_listbox.grid(row=0, column=1, rowspan=4, pady=10, padx=10, sticky='nsew')
        self.category_scrollbar.grid(row=0, column=2, rowspan=4, pady=10, sticky='ns')
        self.add_quiz_button.grid(row=0, column=3, pady=10, padx=10)
        self.edit_quiz_button.grid(row=1, column=3, pady=10, padx=10)

        self.switch_mode_button.grid(row=4, column=0, columnspan=4, pady=10)
        self.question_label.grid(row=5, column=0, columnspan=4, pady=10)
        # self.answer_entry.grid(row=6, column=0, columnspan=4, pady=10)
        self.show_answers_button.grid(row=6, column=0, columnspan=4, pady=10)        
        # self.submit_button.grid(row=8, column=0, columnspan=4)
        # self.next_button.grid(row=8, column=0, columnspan=4, pady=10)
        self.score_label.grid(row=9, column=0, columnspan=4)


    def load_quiz_data(self):
        try:
            if not os.path.isfile(DATA_FILE):
                wb = Workbook()
                ws = wb.active
                ws.title = 'Data'

                ws['A1'] = 'Kanji'
                ws['B1'] = 'Romanji'
                ws['C1'] = 'Significato'
                ws['D1'] = 'Categoria'
                ws['E1'] = 'Tipo (Verbo v /Aggettivo a)'

                for column in ['A', 'B', 'C', 'D', 'E']:
                    ws.column_dimensions[column].width = 30

                wb.save(DATA_FILE)
                self.quiz_data['Generale'] = []
            else:
                wb = load_workbook(DATA_FILE)
                ws = wb.active

                all_quizzes = []

                for row in ws.iter_rows(min_row=2):
                    kanji, romaji, meaning, category, quiz_type = (row[0].value, row[1].value, row[2].value, row[3].value, row[4].value)
                    if quiz_type:
                        quiz_type = quiz_type.lower()
                    if not category or category == "Categoria":
                        category = "Generale"
                    all_quizzes.append({'kanji': kanji, 'romaji': romaji, 'meaning': meaning, 'category': category, 'type': quiz_type})

                # Ordina i quiz in base alla categoria
                all_quizzes.sort(key=lambda x: x['category'])

                # Cancella tutte le righe esistenti (a partire dalla seconda riga)
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.value = None

                # Sovrascrivi le righe nel file Excel con i dati ordinati
                for idx, quiz in enumerate(all_quizzes, start=2):
                    ws.cell(row=idx, column=1, value=quiz['kanji'])
                    ws.cell(row=idx, column=2, value=quiz['romaji'])
                    ws.cell(row=idx, column=3, value=quiz['meaning'])
                    ws.cell(row=idx, column=4, value=quiz['category'])
                    ws.cell(row=idx, column=5, value=quiz['type'])

                # Salva le modifiche nel file Excel
                wb.save(DATA_FILE)

                # Popola il dizionario self.quiz_data con i dati ordinati
                self.quiz_data = {}
                for quiz in all_quizzes:
                    category = quiz['category']
                    if category not in self.quiz_data:
                        self.quiz_data[category] = []
                    self.quiz_data[category].append(quiz)

                # Aggiorna la lista delle categorie
                self.quiz_categories = list(self.quiz_data.keys())

                # Aggiorna la variabile della categoria
                self.category_var.set(self.quiz_categories)                
        except Exception as e:
            messagebox.showerror('Errore', f"Errore durante il caricamento dei dati del quiz: {str(e)}")



    def save_quiz_data(self):
        try:
            # Controlla se il file esiste già
            if os.path.isfile(DATA_FILE):
                # Se esiste, carica il foglio di lavoro
                wb = load_workbook(DATA_FILE)
            else:
                # Se non esiste, crea un nuovo foglio di lavoro
                wb = Workbook()

            # Usa un foglio chiamato "QuizData" o creane uno se non esiste
            if "QuizData" in wb.sheetnames:
                ws = wb["QuizData"]
            else:
                ws = wb.create_sheet(title="QuizData")

            # Aggiungi l'intestazione delle colonne solo se il foglio è vuoto
            if ws.max_row == 1:
                ws.append(['Kanji', 'Romaji', 'Significato', 'Categoria', 'Tipo'])

            # Sovrascrivi le righe con i nuovi dati del quiz
            row_num = 2  # Inizia dalla seconda riga poiché la prima riga ha l'intestazione
            for category, quizzes in self.quiz_data.items():
                for quiz in quizzes:
                    # Se la categoria è vuota o non definita, assegna "Generale"
                    if not category:
                        category = "Generale"
                    for col_num, value in enumerate([quiz['kanji'], quiz['romaji'], quiz['meaning'], category, quiz['type']], start=1):
                        ws.cell(row=row_num, column=col_num, value=value)
                    row_num += 1

            # Cancella le righe in eccesso, se presenti
            for row in ws.iter_rows(min_row=row_num, max_col=5, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

            wb.save(DATA_FILE)

        except PermissionError:
            messagebox.showerror("Errore di salvataggio", "Impossibile salvare i dati del quiz. "
                                                          "Assicurati che il file non sia aperto in un altro programma e riprova."
                                                          "Se il problema persiste, assicurati che il file non sia in uso da un altro programma.")
        except Exception as e:
            messagebox.showerror("Errore di salvataggio", f"Si è verificato un errore durante il salvataggio dei dati del quiz: {str(e)}")


    #Quiz Flow Control 
    ##Queste funzioni gestiscono il flusso del quiz, come caricare nuove domande e controllare le risposte.


    def load_quiz(self, event):
        selected_categories = [self.quiz_categories[i] for i in self.category_listbox.curselection()]
        if selected_categories:
            self.selected_categories = selected_categories
            self.next_question()


    def next_random_quiz(self):
        selected_categories = self.category_listbox.curselection()
        if not selected_categories:
            messagebox.showerror('Errore', 'Seleziona una o più categorie.')
            return

        self.current_category = self.quiz_categories[random.choice(selected_categories)]
    
        if self.current_category not in self.quiz_data:
            self.quiz_data[self.current_category] = []

        if self.current_category not in self.shown_quizzes:
            self.shown_quizzes[self.current_category] = set()

        if len(self.shown_quizzes[self.current_category]) == len(self.quiz_data[self.current_category]):
            # If we have shown all quizzes in this category, reset the set
            self.shown_quizzes[self.current_category] = set()

        remaining_quizzes = [q for q in self.quiz_data[self.current_category] if str(q) not in self.shown_quizzes[self.current_category]]

        if remaining_quizzes:
            quiz = random.choice(remaining_quizzes)
            self.shown_quizzes[self.current_category].add(str(quiz))
            return quiz
        else:
            messagebox.showinfo('No Quizzes', 'La categoria selezionata non contiene ancora quiz.')
            return None


    def next_question(self):
        next_quiz = self.next_random_quiz()
        if next_quiz is not None:
            self.current_quiz = next_quiz

            # Funzione per ottenere il tipo (verbo/aggettivo) se presente
            def get_type(quiz):
                return f" ({quiz['type']})" if quiz['type'] else ""

            # Determina se la risposta corretta è in formato kanji o romaji
            is_kanji_format = bool(next_quiz['kanji'])

            # Ottieni tutte le possibili risposte errate escludendo la risposta corretta, che hanno lo stesso tipo e lo stesso formato
            possible_wrong_answers = [quiz for quiz in self.quiz_data[self.current_category] if quiz != self.current_quiz and quiz.get('type') == self.current_quiz.get('type') and bool(quiz['kanji']) == is_kanji_format]

            # Se non ci sono abbastanza risposte errate dello stesso tipo e formato nella categoria corrente, prendi da tutte le categorie
            if len(possible_wrong_answers) < 2:
                all_other_quizzes = [quiz for cat, quizzes in self.quiz_data.items() for quiz in quizzes if cat != self.current_category and quiz != self.current_quiz and quiz.get('type') == self.current_quiz.get('type') and bool(quiz['kanji']) == is_kanji_format]
                wrong_answers = random.sample(all_other_quizzes, 2 - len(possible_wrong_answers))
                wrong_answers.extend(possible_wrong_answers)
            else:
                wrong_answers = random.sample(possible_wrong_answers, 2)

            if self.quiz_direction == 'kanji to meaning':
                # Creare le opzioni di risposta
                options = [next_quiz['meaning'] + get_type(next_quiz), 
                           wrong_answers[0]['meaning'] + get_type(wrong_answers[0]), 
                           wrong_answers[1]['meaning'] + get_type(wrong_answers[1])]
    
                random.shuffle(options)
                self.option1_button['text'] = options[0]
                self.option2_button['text'] = options[1]
                self.option3_button['text'] = options[2]
                # Nascondi le risposte
                self.option1_button.grid_remove()
                self.option2_button.grid_remove()
                self.option3_button.grid_remove()

                question_type = get_type(next_quiz)
                if next_quiz['kanji']:
                    self.question_label['text'] = f"Quale è il significato di questo kanji/katakana: {next_quiz['kanji']} ({next_quiz['romaji']}){question_type}?"
                else:
                    self.question_label['text'] = f"Quale è il significato di questo romaji: {next_quiz['romaji']}{question_type}?"
    
            else:
                # Creare le opzioni di risposta
                options = [next_quiz['kanji'] + get_type(next_quiz) if next_quiz['kanji'] else next_quiz['romaji'] + get_type(next_quiz),
                           wrong_answers[0]['kanji'] + get_type(wrong_answers[0]) if wrong_answers[0]['kanji'] else wrong_answers[0]['romaji'] + get_type(wrong_answers[0]),
                           wrong_answers[1]['kanji'] + get_type(wrong_answers[1]) if wrong_answers[1]['kanji'] else wrong_answers[1]['romaji'] + get_type(wrong_answers[1])]
    
                random.shuffle(options)
                self.option1_button['text'] = options[0]
                self.option2_button['text'] = options[1]
                self.option3_button['text'] = options[2]
                # Nascondi le risposte
                self.option1_button.grid_remove()
                self.option2_button.grid_remove()
                self.option3_button.grid_remove()
                question_type = get_type(next_quiz)
                if next_quiz['kanji']:
                    self.question_label['text'] = f"Quale kanji/katakana rappresenta questo significato: {next_quiz['meaning']}{question_type}?"
                else:
                    self.question_label['text'] = f"Quale romaji rappresenta questo significato: {next_quiz['meaning']}{question_type}?"
                
            self.next_button['state'] = tk.DISABLED
            self.submit_button['state'] = tk.NORMAL

    def show_answers(self):
        self.option1_button.grid(row=7, column=0, pady=20)
        self.option2_button.grid(row=7, column=1, pady=20)
        self.option3_button.grid(row=7, column=2, pady=20)

    def check_answer(self, selected_option):
        if self.quiz_direction == 'kanji to meaning':
            correct_answer = self.current_quiz['meaning']
        else:
            correct_answer = self.current_quiz['kanji'] if self.current_quiz['kanji'] else self.current_quiz['romaji']

        # Aggiungi il tipo (a/v) alla risposta corretta se presente
        if self.current_quiz.get('type'):
            correct_answer += f" ({self.current_quiz['type']})"

        selected_answer = ""
        if selected_option == 1:
            selected_answer = self.option1_button['text']
        elif selected_option == 2:
            selected_answer = self.option2_button['text']
        elif selected_option == 3:
            selected_answer = self.option3_button['text']

        if selected_answer.strip().lower() == correct_answer.strip().lower():
            # Risposta corretta
            self.correct_answers += 1
        else:
            # Risposta errata
            messagebox.showinfo('Risposta Errata', f'La risposta corretta era: {correct_answer}')

        self.total_questions += 1
        self.update_score()
        self.next_question()
        self.next_button['state'] = tk.NORMAL

        
    def update_score(self):
        self.score_label['text'] = f"Punteggio: {self.correct_answers}/{self.total_questions}"


    def switch_mode(self):
        if self.quiz_direction == 'kanji to meaning':
            self.quiz_direction = 'meaning to kanji'
            messagebox.showinfo('Modalità cambiata', 'Ora devi indovinare il kanji/Katakana dal significato.')
        else:
            self.quiz_direction = 'kanji to meaning'
            messagebox.showinfo('Modalità cambiata', 'Ora devi indovinare il significato dal kanji/katakana.')
        self.next_question()  # Mostra un nuovo quiz dopo il cambio di modalita


    #Quiz Management (Add/Edit) 
    ##Queste funzioni gestiscono l'aggiunta e la modifica dei quiz


    def open_add_category_window(self):
        category = simpledialog.askstring("Aggiungi Categoria", "Inserisci il nome della categoria:")
        if category:
            if category not in self.quiz_categories:
                self.quiz_categories.append(category)
                self.category_listbox.insert(tk.END, category)
                self.quiz_data[category] = []
                messagebox.showinfo('Categoria Aggiunta', 'La categoria è stata aggiunta con successo!')
                self.save_quiz_data()
            else:
                messagebox.showerror('Errore', 'La categoria esiste già.')


    def open_edit_category_window(self):
        selected_index = self.category_listbox.curselection()
        if not selected_index:
            messagebox.showerror('Errore', 'Seleziona una categoria da modificare.')
            return

        category_index = selected_index[0]
        old_category_name = self.quiz_categories[category_index]
        new_category_name = simpledialog.askstring("Modifica Categoria", f"Modifica il nome della categoria '{old_category_name}':")
        if new_category_name:
            if new_category_name not in self.quiz_categories:
                # Rinomina la chiave nel dizionario
                self.quiz_data[new_category_name] = self.quiz_data.pop(old_category_name)

                # Aggiorna e ordina la lista self.quiz_categories
                self.quiz_categories[category_index] = new_category_name
                self.quiz_categories.sort()

                # Aggiorna il category_listbox per riflettere l'ordine alfabetico
                self.category_listbox.delete(0, 'end')
                for category in self.quiz_categories:
                    self.category_listbox.insert('end', category)

                # Salva i dati del quiz
                self.save_quiz_data()
            else:
                messagebox.showerror('Errore', 'La categoria esiste già.')



    def open_edit_quiz_window(self):
        selected_index = self.category_listbox.curselection()
        if not selected_index:
            messagebox.showerror('Errore', 'Seleziona una categoria per modificare un quiz.')
            return

        category_index = selected_index[0]
        selected_category = self.quiz_categories[category_index]
        if selected_category not in self.quiz_data:
            messagebox.showinfo('Nessun quiz', 'La categoria selezionata non contiene ancora quiz. Aggiungine uno prima di modificarlo.')
            return

        edit_quiz_window = tk.Toplevel(self.root)
        edit_quiz_window.title('Modifica Quiz')

        quiz_listbox = tk.Listbox(edit_quiz_window, font=('Arial', 12), width=50)
        quiz_listbox.pack(pady=10)

        for quiz in self.quiz_data[selected_category]:
            quiz_listbox.insert(tk.END, f"{quiz['kanji']} | {quiz['romaji']} | {quiz['meaning']} | {quiz['type']}")

        def update_quiz():
            selected_quiz_index = quiz_listbox.curselection()
            if not selected_quiz_index:
                messagebox.showerror('Errore', 'Seleziona un quiz da modificare.')
                return

            quiz_index = selected_quiz_index[0]
            old_quiz = self.quiz_data[selected_category][quiz_index]

            new_kanji = kanji_entry.get()
            new_meaning = meaning_entry.get()
            new_romaji = romaji_entry.get() 
            new_type = type_entry.get().lower()
            if new_type not in ['a', 'v']:
                new_type = None

            if new_kanji and new_meaning and new_romaji:
                self.quiz_data[selected_category][quiz_index] = {'kanji': new_kanji, 'romaji': new_romaji, 'meaning': new_meaning, 'type': new_type}
                self.save_quiz_data()
                edit_quiz_window.destroy()
            else:
                messagebox.showerror('Errore', 'Inserisci il kanji, il romaji e il significato.')

        kanji_label = tk.Label(edit_quiz_window, text='Kanji:', font=('Arial', 14))
        kanji_label.pack()

        kanji_entry = tk.Entry(edit_quiz_window, font=('Arial', 14))
        kanji_entry.pack()

        romaji_label = tk.Label(edit_quiz_window, text='Romaji:', font=('Arial', 14))
        romaji_label.pack(pady=10)

        romaji_entry = tk.Entry(edit_quiz_window, font=('Arial', 14))
        romaji_entry.pack(pady=10)

        meaning_label = tk.Label(edit_quiz_window, text='Significato:', font=('Arial', 14))
        meaning_label.pack()

        meaning_entry = tk.Entry(edit_quiz_window, font=('Arial', 14))
        meaning_entry.pack()

        type_label = tk.Label(edit_quiz_window, text='Tipo (a/v) Aggettivo o Verbo:', font=('Arial', 14))
        type_label.pack(pady=5)

        type_entry = tk.Entry(edit_quiz_window, font=('Arial', 14))
        type_entry.pack(pady=5)

        edit_button = tk.Button(edit_quiz_window, text='Modifica', font=('Arial', 14), command=update_quiz)
        edit_button.pack(pady=10)


    def open_add_quiz_window(self):
        selected_index = self.category_listbox.curselection()
        if not selected_index:
            messagebox.showerror('Errore', 'Seleziona una categoria per aggiungere un quiz.')
            return

        category_index = selected_index[0]
        selected_category = self.quiz_categories[category_index]

        add_quiz_window = tk.Toplevel(self.root)
        add_quiz_window.title('Aggiungi Quiz')

        kanji_label = tk.Label(add_quiz_window, text='Kanji:', font=('Arial', 14))
        kanji_label.pack(pady=5)

        kanji_entry = tk.Entry(add_quiz_window, font=('Arial', 14))
        kanji_entry.pack(pady=5)

        romaji_label = tk.Label(add_quiz_window, text='Romaji:', font=('Arial', 14))
        romaji_label.pack(pady=5)

        romaji_entry = tk.Entry(add_quiz_window, font=('Arial', 14))
        romaji_entry.pack(pady=5)

        meaning_label = tk.Label(add_quiz_window, text='Significato:', font=('Arial', 14))
        meaning_label.pack(pady=5)

        meaning_entry = tk.Entry(add_quiz_window, font=('Arial', 14))
        meaning_entry.pack(pady=5)

        type_label = tk.Label(add_quiz_window, text='Tipo (a/v) Aggettivo o Verbo:', font=('Arial', 14))
        type_label.pack(pady=5)

        type_entry = tk.Entry(add_quiz_window, font=('Arial', 14))
        type_entry.pack(pady=5)

        def add_quiz():
            kanji = kanji_entry.get()
            meaning = meaning_entry.get()
            romaji = romaji_entry.get()
            quiz_type = type_entry.get().lower()
            if quiz_type not in ['a', 'v']:
                quiz_type = None
            if romaji or meaning:
                self.quiz_data[selected_category].append({'kanji': kanji, 'romaji': romaji, 'meaning': meaning, 'category': selected_category, 'type': quiz_type})
                self.save_quiz_data()
                add_quiz_window.destroy()
            else:
                messagebox.showerror('Errore', 'Inserisci sia il kanji che il significato.')

        add_button = tk.Button(add_quiz_window, text='Aggiungi', font=('Arial', 14), command=add_quiz)
        add_button.pack(pady=10)


if __name__ == "__main__":
    root = tk.Tk()
    style = Style(theme="cyborg")
    # Imposta la larghezza fissa e l'altezza variabile
    #root.minsize(width=800, height=0)
    #root.maxsize(width=800, height=2000)  # Puoi impostare un'altezza massima a tuo piacimento
    QuizApp(root)
    root.mainloop()
